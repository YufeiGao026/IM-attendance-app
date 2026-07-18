import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
import random
from sqlalchemy import text, create_engine, exc
import time
from functools import wraps
import logging
import calendar
import numpy as np

# ========== 多语言支持 ==========

LANGUAGES = {
    "zh": "中文",
    "pt": "Português"
}

TRANSLATIONS = {
    # ---------- 通用 ----------
    "app_title": {"zh": "仓库出勤统计", "pt": "Estatística de Frequência do Armazém"},
    "login_title": {"zh": "🔐 登录", "pt": "🔐 Login"},
    "username": {"zh": "用户名", "pt": "Usuário"},
    "password": {"zh": "密码", "pt": "Senha"},
    "login_button": {"zh": "登录", "pt": "Entrar"},
    "login_error": {"zh": "❌ 用户名或密码错误", "pt": "❌ Usuário ou senha incorretos"},
    "logout": {"zh": "登出", "pt": "Sair"},
    "user_role_admin": {"zh": "管理员", "pt": "Administrador"},
    "user_role_user": {"zh": "普通用户", "pt": "Usuário comum"},
    "language_selector": {"zh": "🌐 语言 / Idioma", "pt": "🌐 Idioma / 语言"},
    "bottom_info": {"zh": "登录用户: {user} | 角色: {role} | 数据模式: 云端 TiDB", "pt": "Usuário: {user} | Função: {role} | Modo de dados: TiDB Cloud"},

    # ---------- 列名 ----------
    "col_region": {"zh": "区域", "pt": "Região"},
    "col_warehouse": {"zh": "仓库", "pt": "Armazém"},
    "col_warehouse_name": {"zh": "仓库", "pt": "Nome do Armazém"},
    "col_date": {"zh": "日期", "pt": "Data"},
    "col_supplier": {"zh": "供应商", "pt": "Fornecedor"},
    "col_shift": {"zh": "班次", "pt": "Turno"},
    "col_worker_type": {"zh": "人员类型", "pt": "Tipo de Mão de Obra"},
    "col_people": {"zh": "人数", "pt": "Nº de Pessoas"},
    "col_volume": {"zh": "操作量", "pt": "Volume"},
    "col_unit_price": {"zh": "单价", "pt": "Preço Unitário"},
    "col_effective_date": {"zh": "生效时间", "pt": "Data de Vigência"},
    "col_expiry_date": {"zh": "失效时间", "pt": "Data de Expiração"},
    "col_sunday": {"zh": "周日_非周日", "pt": "Domingo/Não-Domingo"},
    "col_station": {"zh": "站点", "pt": "Estação"},
    "col_efficiency": {"zh": "人效", "pt": "Eficiência"},
    "col_daily_volume": {"zh": "日均操作量", "pt": "Volume Médio Diário"},
    "col_daily_headcount": {"zh": "日均出勤人数", "pt": "Frequência Média Diária"},
    "col_total_person_days": {"zh": "总出勤人天", "pt": "Total de Dias-pessoa"},
    "col_unit_cost": {"zh": "单人天成本", "pt": "Custo por Pessoa-dia"},
    "col_days": {"zh": "天数", "pt": "Dias"},

    # ---------- 人员类型值 ----------
    "worker_long": {"zh": "长期工", "pt": "Contrato"},
    "worker_daily": {"zh": "日结工", "pt": "Diária"},

    # ---------- Tab 标签 ----------
    "tab_upload_attendance": {"zh": "📤 上传出勤数据", "pt": "📤 Carregar Frequência"},
    "tab_overview": {"zh": "📊 数据总览", "pt": "📊 Visão Geral"},
    "tab_efficiency": {"zh": "📈 外劳人效分析看板", "pt": "📈 Painel de Eficiência"},
    "tab_upload_ops": {"zh": "📊 上传操作量", "pt": "📊 Carregar Volume"},
    "tab_price_card": {"zh": "💰 价卡配置", "pt": "💰 Tabela de Preços"},

    # ---------- 上传出勤数据 ----------
    "attendance_title": {"zh": "📤 上传出勤数据", "pt": "📤 Carregar Frequência"},
    "attendance_instructions": {
        "zh": """
        ### 操作说明
        1. 选择 **日期范围** 和 **一个或多个仓库**。
        2. 点击 **"生成表格"**，系统会生成一个可编辑表格：
           - 前两列为 **仓库** 和 **日期**。
           - 其余列为 **供应商 → 人员类型** 。
        3. 在对应格子中填写出勤人数。
        4. 填写完毕后，点击 **"提交数据"**。
        """,
        "pt": """
        ### Instruções
        1. Selecione o **intervalo de datas** e **um ou mais armazéns**.
        2. Clique em **"Gerar Tabela"** para criar uma tabela editável:
           - As duas primeiras colunas são **Armazém** e **Data**.
           - As demais colunas têm cabeçalho de três níveis: **Fornecedor → Tipo de Mão de Obra**.
        3. Preencha o número de funcionários nas células correspondentes.
        4. Após preencher, clique em **"Enviar Dados"**.
        """
    },
    "attendance_start_date": {"zh": "📅 开始日期", "pt": "📅 Data Inicial"},
    "attendance_end_date": {"zh": "📅 结束日期", "pt": "📅 Data Final"},
    "attendance_select_warehouses": {"zh": "🏭 选择仓库（可多选）", "pt": "🏭 Selecionar Armazéns (múltiplos)"},
    "attendance_generate_btn": {"zh": "📋 生成表格", "pt": "📋 Gerar Tabela"},
    "attendance_no_warehouse_warning": {"zh": "⚠️ 请至少选择一个仓库", "pt": "⚠️ Selecione pelo menos um armazém"},
    "attendance_invalid_date_warning": {"zh": "⚠️ 日期范围无效", "pt": "⚠️ Intervalo de datas inválido"},
    "attendance_no_price_config": {"zh": "⚠️ 所选仓库在价卡表中无配置，请先上传价卡", "pt": "⚠️ Armazém selecionado não possui configuração na tabela de preços. Carregue a tabela primeiro."},
    "attendance_table_subheader": {"zh": "📋 出勤数据表格 ({rows} 行 × {cols} 列)", "pt": "📋 Tabela de Frequência ({rows} linhas × {cols} colunas)"},
    "attendance_submit_btn": {"zh": "📤 提交数据", "pt": "📤 Enviar Dados"},
    "attendance_clear_btn": {"zh": "🗑️ 清空表格", "pt": "🗑️ Limpar Tabela"},
    "attendance_no_data_warning": {"zh": "⚠️ 所有数值均为0，没有需要提交的数据", "pt": "⚠️ Todos os valores são 0, não há dados a enviar"},
    "attendance_no_records_warning": {"zh": "⚠️ 没有找到有效数据（大于0）", "pt": "⚠️ Nenhum dado válido encontrado (maior que 0)"},
    "attendance_success": {"zh": "✅ 全部 {count} 条数据上传成功！版本号：{version}", "pt": "✅ {count} registros enviados com sucesso! Versão: {version}"},
    "attendance_error": {"zh": "❌ 上传失败 {count} 条，请检查数据后重试", "pt": "❌ Falha ao enviar {count} registros. Verifique os dados e tente novamente."},
    "attendance_upload_error": {"zh": "❌ 上传出错：{error}", "pt": "❌ Erro ao enviar: {error}"},
    "attendance_download_template_caption": {"zh": "💡 下载 Excel 模板，表头为三层（供应商 → 班次 → 人员类型），无序号列，填写后复制粘贴到线上表格。", "pt": "💡 Baixe o modelo Excel com cabeçalho de três níveis (Fornecedor → Turno → Tipo). Preencha e copie para a tabela online."},
    "attendance_download_btn": {"zh": "📥 下载模板 (Excel)", "pt": "📥 Baixar Modelo (Excel)"},
    "attendance_info_generate": {"zh": "👆 请选择日期范围、仓库，并点击“生成表格”", "pt": "👆 Selecione o intervalo de datas e armazéns, e clique em “Gerar Tabela”"},

    # ---------- 数据总览 ----------
    "overview_title": {"zh": "📊 数据总览", "pt": "📊 Visão Geral dos Dados"},
    "overview_caption": {"zh": "展示每个仓库按天的最新出勤数据，可按日期范围、区域和仓库筛选。", "pt": "Exibe os dados de frequência mais recentes por dia e armazém. Filtre por intervalo de datas, região e armazém."},
    "overview_month": {"zh": "📅 选择年月", "pt": "📅 Selecionar Mês/Ano"},
    "overview_site": {"zh": "🏢 选择站点", "pt": "🏢 Selecionar Estação"},
    "overview_warehouses": {"zh": "仓库数", "pt": "Armazéns"},
    "overview_total_people": {"zh": "总外劳人数", "pt": "Total de Funcionários"},
    "overview_total_records": {"zh": "总记录数", "pt": "Total de Registros"},
    "overview_uploaders": {"zh": "上传人数", "pt": "Uploaders"},
    "overview_warehouse_summary": {"zh": "各仓库汇总", "pt": "Resumo por Armazém"},
    "overview_export": {"zh": "📥 导出数据", "pt": "📥 Exportar Dados"},
    "overview_export_csv": {"zh": "📥 导出当前数据 (CSV)", "pt": "📥 Exportar dados atuais (CSV)"},
    "overview_no_data": {"zh": "📭 暂无数据，请先上传", "pt": "📭 Sem dados, faça o upload primeiro"},
    "overview_daily_summary": {"zh": "📊 每日汇总", "pt": "📊 Resumo Diário"},
    "overview_daily_detail": {"zh": "📋 每日出勤明细", "pt": "📋 Detalhe Diário de Frequência"},
    "overview_days": {"zh": "天数", "pt": "Dias"},

    # ---------- 效率看板 ----------
    "efficiency_title": {"zh": "📈 外劳人效分析看板", "pt": "📈 Painel de Eficiência"},
    "efficiency_caption": {"zh": "核心指标：人效、日均操作量、日均出勤人数、总出勤人天、单人天成本", "pt": "Indicadores principais: Eficiência, Volume Médio Diário, Frequência Média Diária, Total de Dias-pessoa, Custo por Pessoa-dia"},
    "efficiency_filters": {"zh": "🔍 筛选条件", "pt": "🔍 Filtros"},
    "efficiency_month": {"zh": "年月", "pt": "Mês/Ano"},
    "efficiency_region": {"zh": "区域", "pt": "Região"},
    "efficiency_site": {"zh": "站点", "pt": "Estação"},
    "efficiency_no_data": {"zh": "📭 当前无数据，请先上传出勤数据和操作量数据", "pt": "📭 Sem dados. Carregue os dados de frequência e volume primeiro."},
    "efficiency_read_error": {"zh": "❌ 读取数据失败：{error}", "pt": "❌ Falha ao ler dados: {error}"},
    "efficiency_filter_no_data": {"zh": "该筛选条件下无数据", "pt": "Nenhum dado para este filtro"},
    "efficiency_core_metrics": {"zh": "📊 核心指标", "pt": "📊 Indicadores Principais"},
    "efficiency_metric_efficiency": {"zh": "人效", "pt": "Eficiência"},
    "efficiency_metric_daily_volume": {"zh": "日均操作量", "pt": "Volume Médio Diário"},
    "efficiency_metric_daily_headcount": {"zh": "日均出勤人数", "pt": "Frequência Média Diária"},
    "efficiency_metric_total_person_days": {"zh": "总出勤人天", "pt": "Total de Dias-pessoa"},
    "efficiency_metric_unit_cost": {"zh": "单人天成本", "pt": "Custo por Pessoa-dia"},
    "efficiency_metric_unit": {"zh": "📌 人效单位：票/人/天 | 单人天成本单位：R$", "pt": "📌 Unidades: Eficiência: tickets/pessoa/dia | Custo: R$"},
    "efficiency_station_summary": {"zh": "📋 各站点汇总", "pt": "📋 Resumo por Estação"},
    "efficiency_export": {"zh": "📥 导出数据", "pt": "📥 Exportar Dados"},
    "efficiency_export_csv": {"zh": "📥 导出当前汇总 (CSV)", "pt": "📥 Exportar resumo atual (CSV)"},
    "efficiency_daily_detail": {"zh": "📋 每日效率明细", "pt": "📋 Detalhe Diário de Eficiência"},
    "efficiency_daily_total": {"zh": "📊 每日总览", "pt": "📊 Visão Diária"},

    # ---------- 上传操作量 ----------
    "ops_title": {"zh": "📊 操作量数据上传", "pt": "📊 Carregar Volume de Operações"},
    "ops_instructions": {
        "zh": """
        ---
        ### 📋 使用说明
        1. 点击下方 **"下载模板"** 按钮，下载 Excel 模板
        2. 按模板格式填写数据（**列名必须与模板完全一致**）
        3. 填写完成后，点击 **"选择文件"** 上传
        4. 系统将自动记录上传人和上传时间
        5. **同一站点+日期+班次的数据不可重复上传**
        """,
        "pt": """
        ---
        ### 📋 Instruções
        1. Clique no botão **"Baixar Modelo"** abaixo para baixar o modelo Excel.
        2. Preencha os dados conforme o modelo (**os nomes das colunas devem corresponder exatamente**).
        3. Após preencher, clique em **"Selecionar arquivo"** para fazer o upload.
        4. O sistema registrará automaticamente o uploader e a data/hora.
        5. **Dados duplicados para o mesmo armazém+data+turno não são permitidos.**
        """
    },
    "ops_download_template_btn": {"zh": "📥 下载操作量模板 (Excel)", "pt": "📥 Baixar Modelo de Volume (Excel)"},
    "ops_upload_file": {"zh": "📂 上传操作量文件", "pt": "📂 Carregar arquivo de volume"},
    "ops_choose_file": {"zh": "选择 Excel 或 CSV 文件", "pt": "Selecione arquivo Excel ou CSV"},
    "ops_invalid_columns": {"zh": "❌ 列名与模板不一致！", "pt": "❌ Os nomes das colunas não correspondem ao modelo!"},
    "ops_validation_passed": {"zh": "✅ 校验通过！共 {count} 行数据待上传", "pt": "✅ Validação aprovada! {count} linhas prontas para upload"},
    "ops_submit_btn": {"zh": "✅ 确认上传操作量", "pt": "✅ Confirmar upload de volume"},
    "ops_uploading_status": {"zh": "⏳ 正在上传操作量数据...", "pt": "⏳ Enviando dados de volume..."},
    "ops_upload_success_status": {"zh": "✅ 上传完成！成功 {count} 条", "pt": "✅ Upload concluído! {count} registros"},
    "ops_upload_warning_status": {"zh": "⚠️ 上传完成！成功 {count} 条，失败 {fail} 条", "pt": "⚠️ Upload concluído! {count} sucessos, {fail} falhas"},
    "ops_upload_success": {"zh": "🎉 所有操作量数据已成功上传！", "pt": "🎉 Todos os dados de volume foram enviados com sucesso!"},
    "ops_upload_error": {"zh": "❌ 有 {count} 行上传失败，请检查数据后重试", "pt": "❌ {count} registros falharam. Verifique os dados e tente novamente."},
    "ops_read_error": {"zh": "❌ 读取文件失败：{error}", "pt": "❌ Falha ao ler o arquivo: {error}"},
    "ops_overview_title": {"zh": "📊 操作量数据总览", "pt": "📊 Visão Geral do Volume"},
    "ops_overview_caption": {"zh": "展示已上传的操作量数据，可按日期范围、区域和站点筛选", "pt": "Exibe dados de volume já enviados. Filtre por intervalo de datas, região e estação."},
    "ops_no_data": {"zh": "📭 暂无操作量数据，请先上传", "pt": "📭 Sem dados de volume. Faça o upload primeiro."},
    "ops_month": {"zh": "📅 选择年月", "pt": "📅 Selecionar Mês/Ano"},
    "ops_site": {"zh": "选择站点", "pt": "Selecionar Estação"},
    "ops_total_records": {"zh": "总记录数", "pt": "Total de Registros"},
    "ops_total_volume": {"zh": "总操作量", "pt": "Volume Total"},
    "ops_warehouses": {"zh": "站点数", "pt": "Número de Estações"},
    "ops_days": {"zh": "天数", "pt": "Dias"},
    "ops_station_summary": {"zh": "各站点操作量汇总", "pt": "Resumo por Estação"},
    "ops_export": {"zh": "📥 导出操作量数据", "pt": "📥 Exportar dados de volume"},
    "ops_export_csv": {"zh": "📥 导出操作量数据 (CSV)", "pt": "📥 Exportar dados de volume (CSV)"},
    "ops_read_data_error": {"zh": "⚠️ 读取操作量数据失败：{error}", "pt": "⚠️ Falha ao ler dados de volume: {error}"},

    # ---------- 价卡配置 ----------
    "price_title": {"zh": "💰 价卡配置", "pt": "💰 Tabela de Preços"},
    "price_admin_only": {"zh": "仅管理员（Ivy_Gao）可管理价卡配置", "pt": "Apenas o administrador (Ivy_Gao) pode gerenciar a tabela de preços"},
    "price_current_list": {"zh": "📋 当前价卡列表", "pt": "📋 Lista de Preços Atual"},
    "price_no_config": {"zh": "暂无价卡配置，请联系管理员上传", "pt": "Nenhuma tabela de preços configurada. Entre em contato com o administrador."},
    "price_admin_mode": {"zh": "管理员模式：可下载模板、上传价卡（版本控制，全量导入）", "pt": "Modo administrador: Baixe o modelo, carregue a tabela de preços (controle de versão, importação completa)"},
    "price_current_version": {"zh": "📌 当前生效版本：**{version}**", "pt": "📌 Versão atual: **{version}**"},
    "price_no_version": {"zh": "📌 暂无价卡配置，请上传", "pt": "📌 Nenhuma tabela de preços. Faça o upload."},
    "price_download_template": {"zh": "📋 下载价卡模板", "pt": "📋 Baixar Modelo de Preços"},
    "price_download_btn": {"zh": "📥 下载价卡模板 (Excel)", "pt": "📥 Baixar Modelo (Excel)"},
    "price_template_cols": {"zh": "列名：仓库、供应商、班次、长期工_日结工、周日_非周日、单价、生效时间、失效时间", "pt": "Colunas: Armazém, Fornecedor, Turno, Mão de Obra, Domingo/Não-Domingo, Preço Unitário, Data de Vigência, Data de Expiração"},
    "price_upload_instruction": {"zh": "📤 上传价卡配置（自动生成版本）", "pt": "📤 Carregar Tabela de Preços (versão automática)"},
    "price_upload_caption": {"zh": "每次上传将自动生成唯一版本号（时间戳），新版本将自动成为生效价卡。无需手动输入版本号。", "pt": "Cada upload gera automaticamente um número de versão único (timestamp). A nova versão se tornará automaticamente a tabela de preços ativa."},
    "price_choose_file": {"zh": "选择 Excel 或 CSV", "pt": "Selecionar Excel ou CSV"},
    "price_submit_btn": {"zh": "确认上传", "pt": "Confirmar Upload"},
    "price_missing_file": {"zh": "❌ 请选择文件", "pt": "❌ Selecione um arquivo"},
    "price_invalid_columns": {"zh": "❌ 列名与模板不一致！", "pt": "❌ Os nomes das colunas não correspondem ao modelo!"},
    "price_date_error": {"zh": "❌ 日期格式有误：{error}", "pt": "❌ Formato de data inválido: {error}"},
    "price_price_error": {"zh": "❌ 单价必须为数字：{error}", "pt": "❌ O preço unitário deve ser um número: {error}"},
    "price_upload_success": {"zh": "✅ 价卡配置上传成功！新版本号：{version}，共 {count} 条", "pt": "✅ Tabela de preços enviada com sucesso! Nova versão: {version}, {count} registros"},
    "price_upload_error": {"zh": "❌ 上传失败，请重试", "pt": "❌ Falha no upload. Tente novamente."},
    "price_read_error": {"zh": "❌ 读取文件失败：{error}", "pt": "❌ Falha ao ler o arquivo: {error}"},
    "price_current_version_detail": {"zh": "✅ 当前版本：**{version}** | 上传人：{uploader}", "pt": "✅ Versão atual: **{version}** | Uploader: {uploader}"},
    "price_export_current": {"zh": "📥 导出当前版本价卡", "pt": "📥 Exportar tabela de preços atual"},
    "price_export_btn": {"zh": "📥 导出价卡 ({version})", "pt": "📥 Exportar Preços ({version})"},
    "price_history_expander": {"zh": "📜 查看所有历史版本", "pt": "📜 Ver todas as versões anteriores"},
    "price_history_version": {"zh": "版本号", "pt": "Versão"},
    "price_history_upload_time": {"zh": "上传时间", "pt": "Data/Hora do Upload"},
    "price_history_records": {"zh": "记录数", "pt": "Registros"},
    "price_select_version": {"zh": "选择版本查看详情", "pt": "Selecione a versão para ver detalhes"},
    "price_export_version_btn": {"zh": "📥 导出 {version}", "pt": "📥 Exportar {version}"},

    # ---------- 错误/状态 ----------
    "db_init_error": {"zh": "数据库引擎初始化失败: {error}", "pt": "Falha ao inicializar o motor do banco de dados: {error}"},
    "db_retry_warning": {"zh": "数据库连接超时，正在重试 ({attempt}/{max})...", "pt": "Tempo limite de conexão com o banco de dados. Tentando novamente ({attempt}/{max})..."},
    "db_retry_error": {"zh": "重试{max}次后仍失败: {error}", "pt": "Falha após {max} tentativas: {error}"},
    
    # ---------- 界面通用 ----------
    "clear_msg": {"zh": "清除消息", "pt": "Limpar Mensagem"},
    "download_template": {"zh": "📥 下载模板", "pt": "📥 Baixar Modelo"},
    "confirm_upload": {"zh": "✅ 确认上传", "pt": "✅ Confirmar Upload"},
    "upload_success": {"zh": "✅ 上传成功！", "pt": "✅ Upload realizado com sucesso!"},
    "upload_failed": {"zh": "❌ 上传失败", "pt": "❌ Falha no upload"},
    "no_data": {"zh": "📭 暂无数据", "pt": "📭 Sem dados"},
    "filter_conditions": {"zh": "🔍 筛选条件", "pt": "🔍 Filtros"},
    "select_region": {"zh": "选择区域", "pt": "Selecionar Região"},
    "select_warehouse": {"zh": "选择仓库", "pt": "Selecionar Armazém"},
    "select_site": {"zh": "选择站点", "pt": "Selecionar Estação"},
    "select_warehouse_multi": {"zh": "选择仓库（可多选）", "pt": "Selecionar Armazéns (múltiplos)"},
    "all_regions": {"zh": "全部区域", "pt": "Todas as Regiões"},
    "all_sites": {"zh": "全部站点", "pt": "Todas as Estações"},
    "start_date": {"zh": "开始日期", "pt": "Data Inicial"},
    "end_date": {"zh": "结束日期", "pt": "Data Final"},
    "export_data": {"zh": "📥 导出数据", "pt": "📥 Exportar Dados"},
    "export_csv": {"zh": "📥 导出当前数据 (CSV)", "pt": "📥 Exportar dados atuais (CSV)"},
    "total_people": {"zh": "总外劳人数", "pt": "Total de Funcionários"},
    "avg_daily_people": {"zh": "日均外劳人数", "pt": "Frequência Média Diária"},
    "work_days": {"zh": "工作天数", "pt": "Dias Úteis"},
    "total_records": {"zh": "总记录数", "pt": "Total de Registros"},
    "total_volume": {"zh": "总操作量", "pt": "Volume Total"},
    "avg_daily_volume": {"zh": "日均操作量", "pt": "Volume Médio Diário"},
    "sites_count": {"zh": "站点数", "pt": "Número de Estações"},
    "efficiency_detail": {"zh": "📋 人效明细", "pt": "📋 Detalhe de Eficiência"},
    "efficiency_metric_headcount": {"zh": "总出勤人天", "pt": "Total de Dias-pessoa"},
    "efficiency_metric_avg_headcount": {"zh": "日均出勤人数", "pt": "Frequência Média Diária"},
    "efficiency_metric_unit_cost": {"zh": "单人天成本", "pt": "Custo por Pessoa-dia"},
    "efficiency_metric_workdays": {"zh": "工作天数", "pt": "Dias Úteis"},
    "efficiency_metric_unit_label": {"zh": "📌 人效单位：票/人/天 | 单人天成本单位：R$", "pt": "📌 Unidades: Eficiência: tickets/pessoa/dia | Custo: R$"},
    "efficiency_col_region": {"zh": "区域", "pt": "Região"},
    "efficiency_col_warehouse": {"zh": "仓库", "pt": "Armazém"},
    "efficiency_col_date": {"zh": "日期", "pt": "Data"},
    "efficiency_col_efficiency": {"zh": "人效", "pt": "Eficiência"},
    "efficiency_col_unit_cost": {"zh": "单人天成本", "pt": "Custo por Pessoa-dia"},
    "efficiency_col_headcount": {"zh": "出勤人数", "pt": "Nº de Pessoas"},
    "efficiency_col_volume": {"zh": "操作量", "pt": "Volume"},
    "efficiency_col_cost": {"zh": "总成本", "pt": "Custo Total"},
    "efficiency_summary": {"zh": "【汇总】", "pt": "【Resumo】"},
    "efficiency_all_warehouses": {"zh": "全部仓库 ({count} 个)", "pt": "Todos os Armazéns ({count})"},
    "ops_detail": {"zh": "📋 操作量明细", "pt": "📋 Detalhe de Volume"},
    "ops_col_date": {"zh": "日期", "pt": "Data"},
    "ops_col_site": {"zh": "站点", "pt": "Estação"},
    "ops_col_shift": {"zh": "班次", "pt": "Turno"},
    "ops_col_is_conso": {"zh": "是否集包", "pt": "É Consolidado"},
    "ops_col_volume": {"zh": "操作量", "pt": "Volume"}
}

def _t(key, **kwargs):
    lang = st.session_state.get("language", "zh")
    text = TRANSLATIONS.get(key, {}).get(lang, key)
    if kwargs:
        try:
            return text.format(**kwargs)
        except KeyError:
            return text
    return text

if "language" not in st.session_state:
    st.session_state.language = "zh"

# ========== 页面配置 ==========
st.set_page_config(page_title=_t("app_title"), layout="wide")

# ========== 日志配置 ==========
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ========== 硬编码配置 ==========
DB_HOST = "gateway01.us-east-1.prod.aws.tidbcloud.com"
DB_PORT = 4000
DB_USER = "3CPzLRo7kXD4JE5.root"
DB_PASSWORD = "jdJwdm2QejFDajX4"
DB_NAME = "attendance_db"

# ========== 用户管理 ==========
ADMIN_USERS = ["Ivy_Gao"]
USERS = {
    # 管理员（可以查看所有区域）
    "Ivy_Gao": {"password": "IM_AttendanceData_2606", "role": "admin", "region": "ALL"},
    
    # 普通用户（CDC）
    "IM-EmersonSantaRitaCardoso": {"password": "IM-EmersonSantaRitaCardoso-2026", "role": "user", "region": "CDC"},
    "IM-WendyGodoiDeLima": {"password": "IM-WendyGodoiDeLima-2026", "role": "user", "region": "CDC"},
    
    # 普通用户（FM）
    "IM-GabrielHenriqueLimaDosAnjos": {"password": "IM-GabrielHenriqueLimaDosAnjos-2026", "role": "user", "region": "FM"},
    
    # 普通用户（Middle West）
    "IM-WallaceSoares": {"password": "IM-WallaceSoares-2026", "role": "user", "region": "Middle West"},
    
    # 普通用户（North）
    "IM-VictorEmanoel": {"password": "IM-VictorEmanoel-2026", "role": "user", "region": "North"},
    
    # 普通用户（North East）
    "IM-DouglasBrenoSilvaDeOliveira": {"password": "IM-DouglasBrenoSilvaDeOliveira-2026", "role": "user", "region": "North East"},
    
    # 普通用户（North East2）
    "IM-FabioDeAndradeBezerra": {"password": "IM-FabioDeAndradeBezerra-2026", "role": "user", "region": "North East 2"},
    
    # 普通用户（South）
    "IM-LeonardoJasperCrescencio": {"password": "IM-LeonardoJasperCrescencio-2026", "role": "user", "region": "South"},
    "IM-GUILHERMEMARTINSRODRIGUES": {"password": "IM-GUILHERMEMARTINSRODRIGUES-2026", "role": "user", "region": "South"},
    
    # 普通用户（South East）
    "IM-PedroHenriqueJunioFernandesBrum": {"password": "IM-PedroHenriqueJunioFernandesBrum-2026", "role": "user", "region": "South East"},
    "IM-WillieZhiJieZhang_LoaneJustinoDaSilva": {"password": "IM-WillieZhiJieZhang_LoaneJustinoDaSilva-2026", "role": "user", "region": "South East"},
    
    # 普通用户（SP）
    "IM-EmilyCristinaOliveiraSilva": {"password": "IM-EmilyCristinaOliveiraSilva-2026", "role": "user", "region": "SP"},
    "IM-JamesChiaHaoLin": {"password": "IM-JamesChiaHaoLin-2026", "role": "user", "region": "SP"},
    "IM-ThiagoCardosoDosSantosSilva": {"password": "IM-ThiagoCardosoDosSantosSilva-2026", "role": "user", "region": "SP"},
    
    # 普通用户（ALL）
    "IM-张国庆": {"password": "IM-张国庆-2026", "role": "user", "region": "ALL"},
    "IM-段森楠": {"password": "IM-段森楠-2026", "role": "user", "region": "ALL"},
    "IM-杨涛": {"password": "IM-杨涛-2026", "role": "user", "region": "ALL"},
}



STANDARD_COLS = ["区域", "仓库", "日期", "供应商", "班次", "长期工_日结工", "人数"]
PRICE_CARD_COLS = ["仓库", "供应商", "班次", "长期工_日结工", "周日_非周日", "单价", "生效时间", "失效时间"]
OPS_COLS = ["biz_date", "station_code", "station_name", "class_name", "is_conso", "volume"]

REGION_WAREHOUSE_MAPPING = {
    "CDC": ["CDC-GU", "CDC-SP", "Return Centre"],
    "SP": ["RDC-SP1", "RDC-SP2", "RDC-SP4", "RDC-SP5", "DS BUXI", "DS GRUI", "DS JANI", "DS TAMI"],
    "South": ["DS JMSA", "RDC-PR1", "RDC-RS2", "RDC-SC1"],
    "South East": ["RDC-RJ1", "RDC-MG1", "DC-ES2", "DS BHZI"],
    "North East": ["DC-BA3", "DC-PE2", "RDC-BA1"],
    "North East 2": ["DC-PI2"],
    "Middle West": ["DC-DF2", "DC-MT2", "RDC-GO2", "RDC-TO1"],
    "North": ["DC-PA4"],
    "FM": ["PA GUA", "TP IMN", "Drop BRA", "GLP Guarulhos", "PA MELI", "PA NVS", "PA SAT", "TP BAR", "TP IMG"]
}

ALL_WAREHOUSES = []
for region, warehouses in REGION_WAREHOUSE_MAPPING.items():
    for wh in warehouses:
        ALL_WAREHOUSES.append(wh)

REGION_MAPPING = {}
for region, warehouses in REGION_WAREHOUSE_MAPPING.items():
    for wh in warehouses:
        REGION_MAPPING[wh] = region

@st.cache_resource
def init_db_engine():
    try:
        url = f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4"
        engine = create_engine(url, pool_pre_ping=True, pool_recycle=280, pool_size=5, max_overflow=10, connect_args={'connect_timeout': 30})
        return engine
    except Exception as e:
        st.error(_t("db_init_error", error=str(e)))
        raise

def get_db_connection():
    engine = init_db_engine()
    return engine.connect()

def retry_on_db_error(max_retries=3, delay=1):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if "2013" in str(e) or "Lost connection" in str(e):
                        if attempt < max_retries - 1:
                            st.warning(_t("db_retry_warning", attempt=attempt+1, max=max_retries))
                            time.sleep(delay)
                            st.cache_resource.clear()
                            continue
                        else:
                            logger.error(f"重试{max_retries}次后仍失败: {e}")
                            raise
                    else:
                        logger.error(f"数据库错误: {e}")
                        raise
            return None
        return wrapper
    return decorator

def get_latest_attendance():
    engine = init_db_engine()
    with engine.connect() as conn:
        query = """
        SELECT * FROM attendance
        WHERE 版本号 IN (
            SELECT MAX(版本号) 
            FROM attendance 
            GROUP BY 仓库, 日期, 班次
        )
        """
        df = pd.read_sql(text(query), conn)
        return df


def get_operation_data():
    engine = init_db_engine()
    with engine.connect() as conn:
        query = """
        SELECT 
            station_name AS 仓库,
            biz_date AS 日期,
            class_name AS 班次,
            is_conso,
            volume AS 操作量
        FROM operations
        """
        df = pd.read_sql(text(query), conn)
        if "日期" in df.columns:
            df["日期"] = pd.to_datetime(df["日期"]).dt.strftime("%Y-%m-%d")
        return df

def get_price_card_data(version=None):
    engine = init_db_engine()
    with engine.connect() as conn:
        if version:
            query = "SELECT * FROM price_card WHERE 版本号 = :version"
            df = pd.read_sql(text(query), conn, params={"version": version})
        else:
            query = """
                SELECT * FROM price_card 
                WHERE 上传时间 = (SELECT MAX(上传时间) FROM price_card)
            """
            df = pd.read_sql(text(query), conn)
        return df


@retry_on_db_error(max_retries=3)
def save_attendance_to_db(df):
    engine = init_db_engine()
    with engine.begin() as conn:
        from sqlalchemy import text

        # 把所有的 NaN 替换成 None
        df = df.replace({pd.NA: None, np.nan: None})
       
        records = df.to_dict('records')
        batch_size = 500
        total_rows = len(records)
        success_count = 0

        for i in range(0, total_rows, batch_size):
            batch = records[i:i+batch_size]
            sql = text("""
                INSERT INTO attendance (
                    仓库, 日期, 班次,
                    D0长期工, D0临时工,
                    Enfok长期工, Enfok临时工,
                    Blitz长期工, Blitz临时工,
                    Mission长期工, Mission临时工,
                    Brevi长期工, Brevi临时工,
                    Polly长期工, Polly临时工,
                    GNX长期工, GNX临时工,
                    上传人, 上传时间, 版本号
                ) VALUES (
                    :仓库, :日期, :班次,
                    :D0长期工, :D0临时工,
                    :Enfok长期工, :Enfok临时工,
                    :Blitz长期工, :Blitz临时工,
                    :Mission长期工, :Mission临时工,
                    :Brevi长期工, :Brevi临时工,
                    :Polly长期工, :Polly临时工,
                    :GNX长期工, :GNX临时工,
                    :上传人, :上传时间, :版本号
                )
                ON DUPLICATE KEY UPDATE
                    D0长期工 = VALUES(D0长期工),
                    D0临时工 = VALUES(D0临时工),
                    Enfok长期工 = VALUES(Enfok长期工),
                    Enfok临时工 = VALUES(Enfok临时工),
                    Blitz长期工 = VALUES(Blitz长期工),
                    Blitz临时工 = VALUES(Blitz临时工),
                    Mission长期工 = VALUES(Mission长期工),
                    Mission临时工 = VALUES(Mission临时工),
                    Brevi长期工 = VALUES(Brevi长期工),
                    Brevi临时工 = VALUES(Brevi临时工),
                    Polly长期工 = VALUES(Polly长期工),
                    Polly临时工 = VALUES(Polly临时工),
                    GNX长期工 = VALUES(GNX长期工),
                    GNX临时工 = VALUES(GNX临时工),
                    上传人 = VALUES(上传人),
                    上传时间 = VALUES(上传时间),
                    版本号 = VALUES(版本号)
            """)
            conn.execute(sql, batch)
            success_count += len(batch)
        return success_count, 0

def save_operations_to_db(df):
    import streamlit as st
    engine = init_db_engine()
    with engine.begin() as conn:
        from sqlalchemy import text
        records = df.to_dict('records')
        batch_size = 500
        total_rows = len(records)
        
        st.write(f"📥 准备入库 {total_rows} 条数据，分 { (total_rows + batch_size - 1)//batch_size } 批")
        
        for i in range(0, total_rows, batch_size):
            batch = records[i:i+batch_size]
            sql = text("""
                INSERT INTO operations (biz_date, station_code, station_name, class_name, is_conso, volume, 上传人, 上传时间, 版本号)
                VALUES (:biz_date, :station_code, :station_name, :class_name, :is_conso, :volume, :上传人, :上传时间, :版本号)
                ON DUPLICATE KEY UPDATE
                    station_name = VALUES(station_name),
                    is_conso = VALUES(is_conso),
                    volume = VALUES(volume),
                    上传人 = VALUES(上传人),
                    上传时间 = VALUES(上传时间),
                    版本号 = VALUES(版本号)
            """)
            conn.execute(sql, batch)
            st.write(f"  已完成第 {i//batch_size + 1} 批，本批 {len(batch)} 条")
        
        return len(df), 0

@retry_on_db_error(max_retries=3)
def save_price_card_to_db(df):
    engine = init_db_engine()
    with engine.begin() as conn:
        df.to_sql("price_card", conn, if_exists="append", index=False)
    return len(df), 0

def calculate_cost(price_dict, date, warehouse, shift, worker_type, is_sunday_or_holiday, supplier):
    date_obj = pd.to_datetime(date)
    filtered = price_dict[
        (price_dict["仓库"] == warehouse) &
        (price_dict["班次"] == shift) &
        (price_dict["长期工_日结工"] == worker_type) &
        (price_dict["周日_非周日"] == ("周日" if is_sunday_or_holiday else "非周日")) &
        (price_dict["供应商"] == supplier) &
        (date_obj >= pd.to_datetime(price_dict["生效时间"])) &
        (date_obj <= pd.to_datetime(price_dict["失效时间"]))
    ]
    if not filtered.empty:
        return filtered.iloc[0]["单价"]
    else:
        default = 180.0 if worker_type == "长期工" else 154.0
        return default

def get_work_days_for_month(ym_str):
    year, month = map(int, ym_str.split('-'))
    _, num_days = calendar.monthrange(year, month)
    sundays = sum(1 for day in range(1, num_days+1) if datetime(year, month, day).weekday() == 6)
    return num_days - sundays

def generate_efficiency_data(month=None, region=None, warehouse=None):
    engine = init_db_engine()
    with engine.connect() as conn:
        query = """
        WITH latest_att AS (
            SELECT *,
                   ROW_NUMBER() OVER (
                       PARTITION BY 仓库, DATE_FORMAT(日期, '%Y-%m')
                       ORDER BY 版本号 DESC
                   ) AS rn
            FROM attendance
        ),
        filtered_att AS (
            SELECT * FROM latest_att WHERE rn = 1
        )
        SELECT 
            a.区域, a.仓库, a.日期, a.供应商, a.班次, a.长期工_日结工, a.人数,
            COALESCE(o.volume, 0) AS 操作量,
            CASE WHEN a.人数 > 0 THEN COALESCE(o.volume, 0) / a.人数 ELSE 0 END AS 人效,
            DATE_FORMAT(a.日期, '%Y-%m') AS 年月
        FROM filtered_att a
        LEFT JOIN operations o 
            ON o.station_name = a.仓库 
            AND o.biz_date = a.日期 
            AND o.class_name = a.班次
        WHERE 1=1
        """
        params = {}
        if month and month != "全部":
            query += " AND DATE_FORMAT(a.日期, '%Y-%m') = :month"
            params['month'] = month
        if region and region != "全部":
            query += " AND a.区域 = :region"
            params['region'] = region
        if warehouse and warehouse != "全部":
            query += " AND a.仓库 = :warehouse"
            params['warehouse'] = warehouse
        df = pd.read_sql(text(query), conn, params=params)
        return df

def generate_version(warehouse, work_date):
    engine = init_db_engine()
    with engine.connect() as conn:
        # 查询整个表的最大版本号
        query = """
        SELECT MAX(版本号) as max_version FROM attendance
        """
        result = conn.execute(text(query)).fetchone()
        if result and result[0]:
            max_v = result[0]
            try:
                seq = int(max_v.split("V")[1])
                new_seq = seq + 1
            except:
                new_seq = 1
        else:
            new_seq = 1
        date_str = work_date.replace("-", "")
        return f"{date_str}V{new_seq}"

def get_mock_operation_data():
    dates = [(datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(60)]
    warehouses = ["CDC SP", "CDC RJ", "CDC MG", "TP BAR", "TP IMN", "TP IMG", "PA GUA", "Drop BRA", "GLP Guarulhos"]
    data = []
    for wh in warehouses:
        for d in dates:
            volume = random.randint(500, 3500)
            data.append({"仓库": wh, "日期": d, "操作量": volume})
    return pd.DataFrame(data)

def validate_attendance_data(df):
    errors = []
    required = STANDARD_COLS
    if list(df.columns) != required:
        errors.append(f"列名不匹配，期望: {required}，实际: {list(df.columns)}")
        return errors
    try:
        pd.to_datetime(df["日期"])
    except:
        errors.append("日期列格式不正确，请使用 YYYY-MM-DD")
    if not df["人数"].apply(lambda x: isinstance(x, (int, float)) and x >= 0).all():
        errors.append("人数必须为非负整数（可以为0）")
    return errors

# ========== Session State ==========
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "db_connected" not in st.session_state:
    st.session_state.db_connected = False

# ========== 登录界面 ==========
if not st.session_state.logged_in:
    st.title(_t("login_title"))
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        username = st.text_input(_t("username"))
        password = st.text_input(_t("password"), type="password")
        if st.button(_t("login_button"), use_container_width=True):
            if username in USERS and password == USERS[username]["password"]:
                st.session_state.logged_in = True
                st.session_state.user = username
                st.session_state.is_admin = (USERS[username]["role"] == "admin")
                st.session_state.user_region = USERS[username].get("region", "ALL")  # 默认 ALL
                st.rerun()
            else:
                st.error(_t("login_error"))
    st.stop()

user = st.session_state.user
is_admin = st.session_state.get("is_admin", False)

# ========== 侧边栏 ==========
with st.sidebar:
    st.selectbox(
        _t("language_selector"),
        options=list(LANGUAGES.keys()),
        format_func=lambda x: LANGUAGES[x],
        key="language",
        on_change=lambda: st.rerun()
    )
    st.divider()
    role_text = _t("user_role_admin") if is_admin else _t("user_role_user")
    st.write(f"👤 {user} ({role_text})")
    if st.button(_t("logout") + " / Sair"):
        st.session_state.logged_in = False
        st.rerun()

# ========== 主界面（自定义 Tab，支持翻译且不重置） ==========

# 获取当前语言的 Tab 标签列表（根据角色显示）
all_tab_labels = [
    _t("tab_upload_attendance"),
    _t("tab_overview"),
    _t("tab_efficiency"),
    _t("tab_upload_ops"),
    _t("tab_price_card")
]

# 根据用户角色决定显示哪些 Tab
if st.session_state.get("is_admin", False):
    tab_labels = all_tab_labels
else:
    tab_labels = all_tab_labels[:3]

# 初始化会话状态中的选中索引
if "custom_tab_index" not in st.session_state:
    st.session_state.custom_tab_index = 0

# 自定义 Tab 导航（使用按钮）
cols = st.columns(len(tab_labels))
for i, label in enumerate(tab_labels):
    with cols[i]:
        btn_type = "primary" if st.session_state.custom_tab_index == i else "secondary"
        if st.button(label, type=btn_type, use_container_width=True, key=f"tab_{i}"):
            st.session_state.custom_tab_index = i
            st.rerun()

st.divider()

# 根据选中的索引显示对应的 Tab 内容
active_tab = st.session_state.custom_tab_index

# 【权限拦截】普通用户不允许访问 Tab4 和 Tab5
if not st.session_state.get("is_admin", False) and active_tab >= 3:
    st.session_state.custom_tab_index = 0
    st.rerun()


# ===================== Tab 1: 上传出勤数据 =====================
if active_tab == 0:
    st.title(_t("attendance_title"))
    st.markdown(_t("attendance_instructions"))

    # ---------- 固定配置 ----------
    FIXED_SUPPLIERS = ["D0", "Enfok", "Blitz", "Mission", "Brevi", "Polly", "GNX"]
    FIXED_WORKER_TYPES = ["长期工", "临时工"]

    SHIFTS = ["T1", "T2", "T3"]

    # ---------- 显示持久化成功消息 ----------
    if "upload_success_msg" in st.session_state and st.session_state["upload_success_msg"]:
        st.success(st.session_state["upload_success_msg"])
        if st.button(_t("clear_msg"), key="clear_msg_btn"):
            st.session_state["upload_success_msg"] = ""
            st.rerun()

    # ---------- 选择日期、区域、仓库 ----------
    # 获取所有区域列表（按固定顺序）
    region_order = ["CDC", "SP", "FM", "South East","South", "Middle West","North East", "North East 2", "North"]
    all_regions = [r for r in region_order if r in REGION_WAREHOUSE_MAPPING.keys()]

    # ----- 根据用户区域限制可选的仓库 -----
    user_region = st.session_state.get("user_region", "ALL")
    is_admin = st.session_state.get("is_admin", False)

    # 如果用户是管理员或 region = "ALL"，则可见全部仓库
    if is_admin or user_region == "ALL":
        user_allowed_warehouses = ALL_WAREHOUSES
        user_allowed_regions = all_regions
    else:
        # 普通用户只能看到自己区域的仓库
        user_allowed_warehouses = REGION_WAREHOUSE_MAPPING.get(user_region, [])
        user_allowed_regions = [user_region] if user_region in all_regions else []

    # 如果用户没有任何仓库权限，给出提示
    if not user_allowed_warehouses:
        st.warning("您没有配置任何区域的仓库权限，请联系管理员。")
        st.stop()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        start_date = st.date_input(_t("attendance_start_date"), value=datetime.now().replace(day=1), key="attendance_start_date")
    with col2:
        end_date = st.date_input(_t("attendance_end_date"), value=datetime.now(), key="attendance_end_date")
    with col3:
        # 区域选择框只显示用户有权限的区域
        region_options = [_t("all_regions")] + user_allowed_regions
        selected_region = st.selectbox(
            _t("select_region"),
            region_options,
            key="attendance_region"
        )
    with col4:
        # 根据所选区域动态更新仓库列表（同时受用户权限限制）
        if selected_region == _t("all_regions"):
            available_warehouses = user_allowed_warehouses
            default_warehouses = user_allowed_warehouses[:1] if user_allowed_warehouses else []
        else:
            # 只显示该区域下且在用户权限范围内的仓库（交集）
            region_warehouses = REGION_WAREHOUSE_MAPPING.get(selected_region, [])
            available_warehouses = [wh for wh in region_warehouses if wh in user_allowed_warehouses]
            default_warehouses = available_warehouses  # 默认全选
        
        selected_warehouses = st.multiselect(
            _t("attendance_select_warehouses"),
            available_warehouses,
            default=default_warehouses,
            key="attendance_warehouses"
        )
    # ---------- 生成表格 ----------
    if st.button(_t("attendance_generate_btn"), use_container_width=True, key="generate_btn"):
        st.session_state["upload_success_msg"] = ""
        if not selected_warehouses:
            st.warning(_t("attendance_no_warehouse_warning"))
        else:
            dates = pd.date_range(start=start_date, end=end_date, freq='D').strftime("%Y-%m-%d").tolist()
            if not dates:
                st.warning(_t("attendance_invalid_date_warning"))
            else:
                columns = ["仓库", "日期", "班次"]
                for supplier in FIXED_SUPPLIERS:
                    for worker in FIXED_WORKER_TYPES:
                        columns.append(f"{supplier}{worker}")

                rows = []
                for wh in selected_warehouses:
                    for d in dates:
                        for shift in SHIFTS:
                            row = [wh, d, shift] + [None] * (len(FIXED_SUPPLIERS) * len(FIXED_WORKER_TYPES))
                            rows.append(row)

                df_template = pd.DataFrame(rows, columns=columns)
                st.session_state["attendance_data"] = df_template
                st.session_state["attendance_selected"] = {
                    "warehouses": selected_warehouses,
                    "dates": dates,
                    "shifts": SHIFTS,
                    "suppliers": FIXED_SUPPLIERS,
                    "worker_types": FIXED_WORKER_TYPES,
                }
                st.rerun()

    # ---------- 在线表格编辑 ----------
    if "attendance_data" in st.session_state:
        edited_df = st.data_editor(
            st.session_state["attendance_data"],
            use_container_width=True,
            key="attendance_editor",
            num_rows="fixed",
        )

        if not isinstance(edited_df, pd.DataFrame):
            original_columns = st.session_state["attendance_data"].columns.tolist()
            if isinstance(edited_df, dict):
                edited_df = pd.DataFrame(edited_df)
                edited_df = edited_df.reindex(columns=original_columns, fill_value=None)
            else:
                edited_df = pd.DataFrame(edited_df)

        col_submit, col_clear = st.columns([2, 1])
        with col_submit:
            if st.button(_t("attendance_submit_btn"), type="primary", use_container_width=True, key="submit_data_btn"):
                edited_df = st.session_state["attendance_data"]
                
                # 检查是否有数据（至少有一行）
                if len(edited_df) == 0:
                    st.warning(_t("attendance_no_data_warning"))
                else:
                    # 生成版本号
                    first_warehouse = edited_df.iloc[0]["仓库"]
                    today = datetime.now().strftime("%Y-%m-%d")
                    version = generate_version(first_warehouse, today)
                    
                    # 添加元数据字段
                    edited_df["上传人"] = user
                    edited_df["上传时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    edited_df["版本号"] = version
                    
                    # 直接写入数据库（无需转换）
                    success_count, fail_count = save_attendance_to_db(edited_df)
                    if fail_count == 0:
                        msg = _t("attendance_success", count=success_count, version=version)
                        st.session_state["upload_success_msg"] = msg
                        st.balloons()
                        st.rerun()
                    else:
                        st.error(_t("attendance_error", count=fail_count))

        with col_clear:
            if st.button(_t("attendance_clear_btn"), use_container_width=True, key="clear_table_btn"):
                if "attendance_data" in st.session_state:
                    selected = st.session_state["attendance_selected"]
                    numeric_cols = st.session_state["attendance_data"].columns[3:]
                    zero_rows = []
                    for wh in selected["warehouses"]:
                        for d in selected["dates"]:
                            for shift in SHIFTS:
                                zero_rows.append([wh, d, shift] + [None] * len(numeric_cols))
                    st.session_state["attendance_data"] = pd.DataFrame(zero_rows, columns=st.session_state["attendance_data"].columns)
                    if "attendance_editor" in st.session_state:
                        del st.session_state["attendance_editor"]
                    st.rerun()

    else:
        st.info(_t("attendance_info_generate"))

    # ---------- 模板下载与批量导入 ----------
    st.divider()
    st.subheader(_t("download_template"))

    # 下载模板
    if selected_warehouses:
        dates = pd.date_range(start=start_date, end=end_date, freq='D').strftime("%Y-%m-%d").tolist()
        if dates:
            columns = ["仓库", "日期", "班次"]
            for supplier in FIXED_SUPPLIERS:
                for worker in FIXED_WORKER_TYPES:
                    columns.append(f"{supplier}{worker}")

            rows = []
            for wh in selected_warehouses:
                for d in dates:
                    for shift in SHIFTS:
                        row = [wh, d, shift] + [None] * (len(FIXED_SUPPLIERS) * len(FIXED_WORKER_TYPES))
                        rows.append(row)

            df_template = pd.DataFrame(rows, columns=columns)

            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Font, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "出勤数据"
            for i, col in enumerate(df_template.columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = 15

            for col_idx, col_name in enumerate(df_template.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_idx, row_data in enumerate(df_template.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else "")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            for row in range(1, len(df_template)+2):
                for col in range(1, len(df_template.columns)+1):
                    ws.cell(row=row, column=col).border = thin_border

            output = BytesIO()
            wb.save(output)
            template_bytes = output.getvalue()

            col_dl, _ = st.columns([1, 3])
            with col_dl:
                st.download_button(
                    label=_t("attendance_download_btn"),
                    data=template_bytes,
                    file_name="出勤模板_完整.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_template_btn"
                )
            st.caption(_t("attendance_download_template_caption"))

    # 批量上传文件
    uploaded_file = st.file_uploader(
        _t("attendance_download_template_caption"),
        type=["xlsx", "xls", "csv"],
        key="batch_upload_attendance"
    )

    if uploaded_file is not None:
        # ===== 持久化显示上次导入结果 =====
        if "attendance_batch_msg" in st.session_state and st.session_state["attendance_batch_msg"]:
            st.success(st.session_state["attendance_batch_msg"])
            if st.button(_t("clear_msg"), key="clear_att_batch_msg"):
                st.session_state["attendance_batch_msg"] = ""
                st.rerun()
    
    

        with st.status("📤 正在校验文件...", expanded=True) as status:
            try:
                # 1. 读取文件
                status.update(label="📖 读取文件...")
                if uploaded_file.name.endswith(".csv"):
                    df_batch = pd.read_csv(uploaded_file)
                else:
                    df_batch = pd.read_excel(uploaded_file, engine="openpyxl")
                status.update(label=f"✅ 读取成功，共 {len(df_batch)} 行")

                # 2. 校验列名（必须完全匹配）
                expected_cols = ["仓库", "日期", "班次"] + [f"{s}{w}" for s in FIXED_SUPPLIERS for w in FIXED_WORKER_TYPES]
                if list(df_batch.columns) != expected_cols:
                    st.error(_t("attendance_invalid_date_warning"))  # 复用，实际应改为列名错误
                    status.update(label="❌ 列名校验失败", state="error")
                    st.stop()

                # 3. 计算预期行数（所有组合）
                dates = pd.date_range(start=start_date, end=end_date, freq='D').strftime("%Y-%m-%d").tolist()
                expected_rows = len(selected_warehouses) * len(dates) * len(SHIFTS)
                if len(df_batch) != expected_rows:
                    st.error(f"❌ 行数不正确！\n模板应有 {expected_rows} 行（{len(selected_warehouses)} 个仓库 × {len(dates)} 天 × {len(SHIFTS)} 个班次），实际上传 {len(df_batch)} 行。")
                    status.update(label="❌ 行数校验失败", state="error")
                    st.stop()

                # 4. 校验每行的仓库、日期、班次是否合法
                invalid_warehouses = df_batch[~df_batch["仓库"].isin(selected_warehouses)]["仓库"].unique()
                if len(invalid_warehouses) > 0:
                    st.error(f"❌ 存在不在选定列表中的仓库：{invalid_warehouses}")
                    status.update(label="❌ 仓库校验失败", state="error")
                    st.stop()

                try:
                    df_batch["日期"] = pd.to_datetime(df_batch["日期"]).dt.strftime("%Y-%m-%d")
                except Exception as e:
                    st.error(f"❌ 日期格式错误：{e}")
                    status.update(label="❌ 日期解析失败", state="error")
                    st.stop()
                invalid_dates = df_batch[(df_batch["日期"] < start_date.strftime("%Y-%m-%d")) | 
                                        (df_batch["日期"] > end_date.strftime("%Y-%m-%d"))]["日期"].unique()
                if len(invalid_dates) > 0:
                    st.error(f"❌ 存在不在选定日期范围内的日期：{invalid_dates}")
                    status.update(label="❌ 日期范围校验失败", state="error")
                    st.stop()

                invalid_shifts = df_batch[~df_batch["班次"].isin(SHIFTS)]["班次"].unique()
                if len(invalid_shifts) > 0:
                    st.error(f"❌ 存在不允许的班次：{invalid_shifts}，只允许 {SHIFTS}")
                    status.update(label="❌ 班次校验失败", state="error")
                    st.stop()

                # 5. 解析有效数据（提取人数 >0 的记录）
                status.update(label="📊 解析数据记录...")
                records = []
                numeric_cols = df_batch.columns[3:]
                for idx, row in df_batch.iterrows():
                    warehouse = row["仓库"]
                    date = row["日期"]
                    shift = row["班次"]
                    
                    # 检查仓库、日期、班次是否有内容（非空）
                    if pd.isna(warehouse) or pd.isna(date) or pd.isna(shift):
                        continue  # 这三列任意为空，跳过该行
                    
                    # 对该行的每个供应商×人员类型列生成记录
                    for col in numeric_cols:
                        val = row[col]
                        if pd.isna(val):
                            val = 0  # 空值视为0
                        # 解析列名
                        if col.endswith("长期工"):
                            supplier = col[:-3]
                            worker_type_db = "长期工"
                        elif col.endswith("临时工"):
                            supplier = col[:-3]
                            worker_type_db = "日结工"
                        else:
                            continue
                        region = REGION_MAPPING.get(warehouse, "未知")
                        records.append({
                            "区域": region,
                            "仓库": warehouse,
                            "日期": date,
                            "班次": shift,
                            "供应商": supplier,
                            "长期工_日结工": worker_type_db,
                            "人数": int(val),
                        })

                if not records:
                    st.warning(_t("attendance_no_data_warning"))
                    status.update(label="⚠️ 无有效数据", state="error")
                    st.stop()

                st.info(f"📊 数据校验完成，共 {len(df_batch)} 行数据，请点击确认上传")
                status.update(label=f"✅ 数据校验完成，共 {len(df_batch)} 行数据", state="complete")

                # 将解析结果暂存到 session_state，供确认按钮使用
                st.session_state["batch_records"] = records
                st.session_state["batch_df"] = df_batch

            except Exception as e:
                st.error(f"❌ 处理文件时出错：{e}")
                status.update(label=f"❌ 处理失败：{e}", state="error")
                st.stop()

        # 校验通过后，显示确认按钮（放在 status 外部）
        if "batch_records" in st.session_state and st.session_state["batch_records"]:
            if st.button(_t("confirm_upload"), key="batch_submit"):
                with st.status("📤 正在导入数据...", expanded=True) as import_status:
                    try:
                        # 获取已经校验过的宽表数据（不包含元数据）
                        df_to_submit = st.session_state["batch_df"].copy()
                        
                        # 生成版本号（使用第一个仓库）
                        first_warehouse = df_to_submit["仓库"].iloc[0]
                        today = datetime.now().strftime("%Y-%m-%d")
                        version = generate_version(first_warehouse, today)
                        
                        # 添加元数据字段
                        df_to_submit["上传人"] = user
                        df_to_submit["上传时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        df_to_submit["版本号"] = version
                        
                        # 直接插入宽表（无需展开）
                        success_count, fail_count = save_attendance_to_db(df_to_submit)
                        
                        if fail_count == 0:
                            msg = _t("attendance_success", count=success_count, version=version)
                            st.session_state["attendance_batch_msg"] = msg
                            st.success(msg)
                            st.balloons()
                            import_status.update(label=msg, state="complete")
                            del st.session_state["batch_records"]
                            st.rerun()
                        else:
                            error_msg = _t("attendance_error", count=fail_count)
                            st.session_state["attendance_batch_msg"] = error_msg
                            st.error(error_msg)
                            import_status.update(label=f"❌ 导入失败 {fail_count} 条", state="error")
                    except Exception as e:
                        error_msg = _t("attendance_upload_error", error=str(e))
                        st.session_state["attendance_batch_msg"] = error_msg
                        st.error(error_msg)
                        import_status.update(label=f"❌ 导入出错：{e}", state="error")

# ===================== Tab 2: 数据总览 =====================
elif active_tab == 1:
    st.title(_t("overview_title"))
    st.caption(_t("overview_caption"))

    # ---------- 获取数据 ----------
    df_raw = get_latest_attendance()  # 修正：函数无参数
    if len(df_raw) == 0:
        st.info(_t("overview_no_data"))
    else:
        # 日期列处理
        df_raw["日期"] = pd.to_datetime(df_raw["日期"])
        
        # 获取所有仓库列表（来自数据）
        all_warehouses = sorted(df_raw["仓库"].unique())
        
        # 区域顺序（与 Tab 1 保持一致）
        region_order = ["CDC", "SP", "South", "South East", "North East", "North East 2", "Middle West", "North", "FM"]
        all_regions = [r for r in region_order if r in REGION_WAREHOUSE_MAPPING.keys()]

        # ----- 根据用户区域限制可选的仓库 -----
        user_region = st.session_state.get("user_region", "ALL")
        is_admin = st.session_state.get("is_admin", False)

        # 如果用户是管理员或 region = "ALL"，则可见全部仓库
        if is_admin or user_region == "ALL":
            user_allowed_warehouses = all_warehouses
            user_allowed_regions = all_regions
        else:
            # 普通用户只能看到自己区域的仓库（交集：区域映射中的仓库且数据中存在）
            region_warehouses = REGION_WAREHOUSE_MAPPING.get(user_region, [])
            user_allowed_warehouses = [wh for wh in region_warehouses if wh in all_warehouses]
            user_allowed_regions = [user_region] if user_region in all_regions else []

        # 如果用户没有任何仓库权限，给出提示
        if not user_allowed_warehouses:
            st.warning("您没有配置任何区域的仓库权限，请联系管理员。")
            st.stop()
        
        # ---------- 筛选栏 ----------
        st.subheader(_t("filter_conditions"))
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        with col_f1:
            start_date = st.date_input(
                _t("start_date"),
                value=df_raw["日期"].min().date(),
                key="overview_start_date"
            )
        with col_f2:
            end_date = st.date_input(
                _t("end_date"),
                value=df_raw["日期"].max().date(),
                key="overview_end_date"
            )
        with col_f3:
            # 区域选择框只显示用户有权限的区域
            region_options = [_t("all_regions")] + user_allowed_regions
            selected_region = st.selectbox(
                _t("select_region"),
                region_options,
                key="overview_region"
            )
        with col_f4:
            # 根据所选区域动态更新仓库列表（同时受用户权限限制）
            if selected_region == _t("all_regions"):
                available_warehouses = user_allowed_warehouses
                default_warehouses = available_warehouses
            else:
                region_warehouses = REGION_WAREHOUSE_MAPPING.get(selected_region, [])
                available_warehouses = [wh for wh in region_warehouses if wh in user_allowed_warehouses]
                default_warehouses = available_warehouses
            
            selected_warehouses = st.multiselect(
                _t("select_warehouse_multi"),
                options=available_warehouses,
                default=default_warehouses,
                key="overview_warehouses"
            )
        
        # ---------- 数据过滤 ----------
        df_filtered = df_raw.copy()
        df_filtered = df_filtered[
            (df_filtered["日期"] >= pd.to_datetime(start_date)) & 
            (df_filtered["日期"] <= pd.to_datetime(end_date))
        ]
        # 区域过滤（根据选择的仓库列表筛选）
        if selected_warehouses:
            df_filtered = df_filtered[df_filtered["仓库"].isin(selected_warehouses)]
        
        if len(df_filtered) == 0:
            st.warning(_t("efficiency_filter_no_data"))
        else:
            # ---------- 统计指标 ----------
            total_records = len(df_filtered)
            total_warehouses = df_filtered["仓库"].nunique()
            # 计算所有供应商列的总和
            people_cols = [col for col in df_filtered.columns if col.endswith("长期工") or col.endswith("临时工")]
            total_people = int(df_filtered[people_cols].sum().sum()) if people_cols else 0
            
            # ---------- 计算工作天数 ----------
            date_range = pd.date_range(start=start_date, end=end_date, freq='D')
            work_days = sum(1 for d in date_range if d.weekday() != 6)  # 非周日天数

            avg_daily_people = total_people / work_days if work_days > 0 else 0

            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric(_t("total_people"), f"{total_people:,}")
            col2.metric(_t("avg_daily_people"), f"{avg_daily_people:.1f}")
            col3.metric(_t("work_days"), f"{work_days}")
            col4.metric(_t("overview_warehouses"), total_warehouses)
            col5.metric(_t("total_records"), total_records)
            
            st.divider()
            
            # ---------- 展示明细数据 ----------
            st.subheader(_t("overview_daily_detail"))
            
            # 重命名列，使显示更友好
            rename_map = {
                "仓库": _t("col_warehouse_name"),
                "日期": _t("col_date"),
                "班次": _t("col_shift"),
                "上传人": "上传人",
                "上传时间": "上传时间",
                "版本号": "版本号"
            }
            # 如果有区域列，也重命名
            if "区域" in df_filtered.columns:
                rename_map["区域"] = _t("col_region")
            display_df = df_filtered.rename(columns=rename_map)
            
            # 格式化日期列
            if _t("col_date") in display_df.columns:
                display_df[_t("col_date")] = pd.to_datetime(display_df[_t("col_date")]).dt.strftime("%Y-%m-%d")
            
            # 确定列顺序：区域、仓库、日期、班次 + 所有供应商列 + 上传人、上传时间、版本号
            fixed_cols = []
            if _t("col_region") in display_df.columns:
                fixed_cols.append(_t("col_region"))
            fixed_cols.extend([_t("col_warehouse_name"), _t("col_date"), _t("col_shift")])
            
            supplier_cols = [col for col in display_df.columns if col.endswith("长期工") or col.endswith("临时工")]
            # 按固定顺序排列供应商列
            fixed_suppliers = ["D0", "Enfok", "Blitz", "Mission", "Brevi", "Polly", "GNX"]
            fixed_worker_types = ["长期工", "临时工"]
            ordered_supplier_cols = []
            for supplier in fixed_suppliers:
                for worker in fixed_worker_types:
                    col_name = f"{supplier}{worker}"
                    if col_name in display_df.columns:
                        ordered_supplier_cols.append(col_name)
            
            final_cols = fixed_cols + ordered_supplier_cols + ["上传人", "上传时间", "版本号"]
            final_cols = [col for col in final_cols if col in display_df.columns]
            display_df = display_df[final_cols]
            
            st.dataframe(display_df, use_container_width=True)
            
            # ---------- 导出功能 ----------
            st.subheader(_t("overview_export"))
            csv = display_df.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                label=_t("overview_export_csv"),
                data=csv,
                file_name=f"仓库出勤汇总_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                key="overview_export_btn"
            )

# ===================== Tab 3: 外劳人效分析看板 =====================
elif active_tab == 2:
    st.title(_t("efficiency_title"))
    st.caption(_t("efficiency_caption"))
    st.markdown("""
    <style>
    div[data-testid="metric-container"] .stMetricValue { font-size: 28px !important; }
    div[data-testid="metric-container"] .stMetricLabel { font-size: 14px !important; }
    div[data-testid="metric-container"] .stMetricDelta { font-size: 12px !important; }
    </style>
    """, unsafe_allow_html=True)

    try:
        # 加载价卡数据并转换为查找字典（一次性）
        price_df = get_price_card_data()
        if not price_df.empty:
            # 构建一个字典，key 是 (仓库, 班次, 人员类型, 是否周日, 供应商)，value 是单价
            price_dict = {}
            for _, row in price_df.iterrows():
                key = (
                    row["仓库"],
                    row["班次"],
                    row["长期工_日结工"],
                    row["周日_非周日"],
                    row["供应商"]
                )
                price_dict[key] = row["单价"]
        else:
            price_dict = {}
        # ---------- 1. 获取出勤数据（宽表） ----------
        
        att_df = get_latest_attendance()
        if len(att_df) == 0:
            st.info(_t("efficiency_no_data"))
            st.stop()
        
        # ---------- 2. 将宽表转成长表 ----------
        
        # 定义固定列和动态列
        id_vars = ["仓库", "日期", "班次", "上传人", "上传时间", "版本号"]
        # 如果有区域列也保留（新宽表可能没有）
        if "区域" in att_df.columns:
            id_vars.insert(0, "区域")
        
        # 找出所有供应商列（以"长期工"或"临时工"结尾的列）
        value_vars = [col for col in att_df.columns if col.endswith("长期工") or col.endswith("临时工")]
        if not value_vars:
            st.warning("未找到人员类型列，请检查出勤数据。")
            st.stop()
        
        # 使用 melt 将宽表转成长表
        long_df = pd.melt(
            att_df,
            id_vars=id_vars,
            value_vars=value_vars,
            var_name="供应商_人员类型",
            value_name="人数"
        )
        
        # 拆分供应商和人员类型
        long_df["供应商"] = long_df["供应商_人员类型"].str.replace("长期工", "").str.replace("临时工", "")
        long_df["长期工_日结工"] = long_df["供应商_人员类型"].str.extract(r'(长期工|临时工)')[0]
        long_df = long_df.drop(columns=["供应商_人员类型"])
        
        # ---------- 3. 获取操作量数据 ----------
        
        ops_df = get_operation_data()  # 返回: 仓库, 日期, 班次, 操作量
        
        # ---------- 4. 按站点+日期汇总操作量 ----------
        # ---------- 4. 按站点+日期汇总操作量 ----------
        # 统一日期格式
        long_df["日期"] = pd.to_datetime(long_df["日期"])
        ops_df["日期"] = pd.to_datetime(ops_df["日期"])

        # 🔑 关键修复：统一仓库名称（去除空格，转大写）
        long_df["仓库"] = long_df["仓库"].astype(str).str.strip().str.upper()
        ops_df["仓库"] = ops_df["仓库"].astype(str).str.strip().str.upper()

        # 先按 仓库+日期 汇总操作量（忽略班次）
        ops_df_grouped = ops_df.groupby(["仓库", "日期"], as_index=False)["操作量"].sum()

        

        

        # 再与出勤数据合并（只匹配仓库+日期）
        merged = pd.merge(
            long_df,
            ops_df_grouped,
            left_on=["仓库", "日期"],
            right_on=["仓库", "日期"],
            how="left"
        )

        # 如果操作量匹配不上，填充0
        if "操作量" in merged.columns:
            merged["操作量"] = merged["操作量"].fillna(0)
        else:
            merged["操作量"] = 0

        # 统计匹配情况
        matched_count = (merged["操作量"] > 0).sum()
        total_count = len(merged)
        


        
        # 计算人效
        merged["人效"] = merged.apply(
            lambda row: row["操作量"] / row["人数"] if row["人数"] > 0 else 0,
            axis=1
        )
        merged["年月"] = pd.to_datetime(merged["日期"]).dt.strftime("%Y-%m")
        
        # 确保日期格式为datetime
        merged["日期"] = pd.to_datetime(merged["日期"])
        
        # ---------- 5. 整理区域映射 ----------
       
        # 如果宽表中没有区域列，从映射表获取
        if "区域" not in merged.columns:
            merged["区域"] = merged["仓库"].map(REGION_MAPPING).fillna("未知")
        
        # ---------- 6. 获取所有区域列表 ----------
     
        region_order = ["CDC", "SP", "South", "South East", "North East", "North East 2", "Middle West", "North", "FM"]
        all_regions = [r for r in region_order if r in REGION_WAREHOUSE_MAPPING.keys()]
        all_sites_in_data = sorted(merged["仓库"].unique())

        # ---------- 7. 筛选条件 ----------
        
        st.subheader(_t("efficiency_filters"))
        
        min_date = merged["日期"].min() if not merged.empty else datetime.now().date()
        max_date = merged["日期"].max() if not merged.empty else datetime.now().date()
        
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        with col_f1:
            start_date = st.date_input(_t("start_date"), value=min_date, key="eff_start_date")
        with col_f2:
            end_date = st.date_input(_t("end_date"), value=max_date, key="eff_end_date")
        with col_f3:
            selected_region = st.selectbox(_t("select_region"), [_t("all_regions")] + all_regions, key="eff_region")
        with col_f4:
            if selected_region == _t("all_regions"):
                available_sites = all_sites_in_data
            else:
                region_sites = REGION_WAREHOUSE_MAPPING.get(selected_region, [])
                available_sites = [site for site in region_sites if site in all_sites_in_data]
            selected_warehouse = st.selectbox(_t("select_site"), [_t("all_sites")] + available_sites, key="eff_site")

        # ---------- 8. 数据过滤 ----------
        filtered_df = merged.copy()
        filtered_df = filtered_df[
            (filtered_df["日期"] >= pd.to_datetime(start_date)) & 
            (filtered_df["日期"] <= pd.to_datetime(end_date))
        ]
        if selected_region != _t("all_regions"):
            region_sites = REGION_WAREHOUSE_MAPPING.get(selected_region, [])
            filtered_df = filtered_df[filtered_df["仓库"].isin(region_sites)]
        if selected_warehouse != _t("all_sites"):
            filtered_df = filtered_df[filtered_df["仓库"] == selected_warehouse]

        if len(filtered_df) == 0:
            st.warning(_t("efficiency_filter_no_data"))
            st.stop()

        # ---------- 9. 计算工作天数 ----------
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        work_days = sum(1 for d in date_range if d.weekday() != 6)

       

        # ---------- 11. 成本计算函数 ----------
        
        def calculate_group_cost(df, price_dict):
            if df.empty:
                return 0.0
            df = df.copy()
            df["is_sunday_or_holiday"] = df["日期"].apply(
                lambda d: (d.weekday() == 6) or (d.month == 5 and d.day == 1)
            )
            df["周日_非周日"] = df["is_sunday_or_holiday"].map({True: "周日", False: "非周日"})
            
            grouped = df.groupby(
                ["仓库", "供应商", "日期", "班次", "长期工_日结工", "周日_非周日"],
                as_index=False
            ).agg({"人数": "sum"})
            grouped["人数"] = grouped["人数"].fillna(0)
            
            # 向量化匹配单价（不再用循环查表）
            def get_unit_price(row):
                key = (
                    row["仓库"],
                    row["班次"],
                    row["长期工_日结工"],
                    row["周日_非周日"],
                    row["供应商"]
                )
                unit_price = price_dict.get(key)
                if unit_price is None:
                    unit_price = 180.0 if row["长期工_日结工"] == "长期工" else 154.0
                return unit_price
            
            grouped["单价"] = grouped.apply(get_unit_price, axis=1)
            grouped["成本"] = grouped["人数"] * grouped["单价"]
            return grouped["成本"].sum()

        # ---------- 12. 计算整体指标 ----------
        total_person_days = int(filtered_df["人数"].sum())
        # 按（仓库, 日期）去重后求和操作量
        daily_ops = filtered_df.groupby(["仓库", "日期"])["操作量"].first().reset_index()
        total_volume = int(daily_ops["操作量"].sum())
        total_cost = calculate_group_cost(filtered_df, price_dict)
        
        avg_daily_headcount = total_person_days / work_days if work_days > 0 else 0
        efficiency = total_volume / total_person_days if total_person_days > 0 else 0
        unit_cost = total_cost / total_person_days if total_person_days > 0 else 0

        # ---------- 13. 核心指标 ----------
        st.subheader(_t("efficiency_core_metrics"))
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric(_t("efficiency_metric_headcount"), f"{total_person_days:,}")
        c2.metric(_t("efficiency_metric_avg_headcount"), f"{avg_daily_headcount:.1f}")
        c3.metric(_t("efficiency_metric_efficiency"), f"{efficiency:.0f}")
        c4.metric(_t("efficiency_metric_unit_cost"), f"R${unit_cost:.2f}")
        c5.metric(_t("efficiency_metric_workdays"), f"{work_days}")
        st.caption(_t("efficiency_metric_unit_label"))

        # ---------- 14. 人效明细 ----------
        st.subheader(_t("efficiency_detail"))

        # 按 日期 + 仓库 分组
        # 按 日期 + 仓库 分组
        date_warehouse_list = filtered_df.groupby(["日期", "仓库"]).agg({
            "区域": "first",
            "人数": "sum",
            "操作量": "max"   # 改为 max，因为同一天同一仓库操作量相同
        }).reset_index()

        detail_rows = []
        for _, row in date_warehouse_list.iterrows():
            date = row["日期"]
            warehouse = row["仓库"]
            region = row["区域"]
            
            # 获取该天该仓库的原始数据用于计算成本
            day_warehouse_df = filtered_df[
                (filtered_df["日期"] == date) & 
                (filtered_df["仓库"] == warehouse)
            ]
            
            day_person_days = int(row["人数"])
            day_volume = int(row["操作量"])
            day_cost = calculate_group_cost(day_warehouse_df, price_dict)
            
            day_efficiency = day_volume / day_person_days if day_person_days > 0 else 0
            day_unit_cost = day_cost / day_person_days if day_person_days > 0 else 0
            day_avg_cost = day_cost / work_days if work_days > 0 else 0
            day_avg_headcount = day_person_days / work_days if work_days > 0 else 0
            day_avg_volume = day_volume / work_days if work_days > 0 else 0
            
            detail_rows.append({
                _t("efficiency_col_region"): region,
                _t("efficiency_col_warehouse"): warehouse,
                _t("efficiency_col_date"): date.strftime("%Y-%m-%d"),
                _t("efficiency_col_efficiency"): f"{day_efficiency:.0f}",
                _t("efficiency_col_unit_cost"): f"R${day_unit_cost:.2f}",
                _t("efficiency_col_headcount"): f"{day_person_days:,}",
                _t("efficiency_col_volume"): f"{day_volume:,}",
                _t("efficiency_col_cost"): f"R${day_cost:,.2f}"
            })

        # 构建明细 DataFrame
        detail_df = pd.DataFrame(detail_rows)

        # ---------- 15. 构建汇总行 ----------
        summary_row = {
            _t("efficiency_col_region"): _t("efficiency_summary"),
            _t("efficiency_col_warehouse"): _t("efficiency_all_warehouses").replace("{count}", str(len(date_warehouse_list['仓库'].unique()))),
            _t("efficiency_col_date"): f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}",
            _t("efficiency_col_efficiency"): f"{efficiency:.0f}",
            _t("efficiency_col_unit_cost"): f"R${unit_cost:.2f}",
            _t("efficiency_col_headcount"): f"{total_person_days:,}",
            _t("efficiency_col_volume"): f"{total_volume:,}",
            _t("efficiency_col_cost"): f"R${total_cost:,.2f}"
        }

        summary_df = pd.DataFrame([summary_row])
        final_detail_df = pd.concat([summary_df, detail_df], ignore_index=True)

        st.dataframe(final_detail_df, use_container_width=True)

        
        # ============================================================
        # ---------- 17. 数据下载 ----------
        # ============================================================
        st.subheader(_t("efficiency_export"))
        csv = final_detail_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            label=_t("efficiency_export_csv"),
            data=csv,
            file_name=f"外劳人效分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            key="eff_export_btn"
        )

    except Exception as e:
        st.error(f"数据加载错误：{e}")
        import traceback
        st.code(traceback.format_exc())

# ===================== Tab 4: 上传操作量 =====================
elif active_tab == 3:
    st.title(_t("ops_title"))
    st.markdown(_t("ops_instructions"))
    
    st.subheader("📋 " + _t("ops_download_template_btn").replace("📥 ", ""))
    template_ops_df = pd.DataFrame(columns=OPS_COLS)
    ops_output = BytesIO()
    with pd.ExcelWriter(ops_output, engine="openpyxl") as writer:
        template_ops_df.to_excel(writer, index=False, sheet_name="操作量数据")
    ops_template_bytes = ops_output.getvalue()
    col_ops_download, _ = st.columns([1, 3])
    with col_ops_download:
        st.download_button(
            label=_t("ops_download_template_btn"),
            data=ops_template_bytes,
            file_name="操作量模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="ops_download_btn"
        )
    st.caption("列名：biz_date、station_code、station_name、class_name、is_conso、volume（is_conso 为 0 或 1）")
    st.divider()
    st.subheader(_t("ops_upload_file"))
    ops_uploaded_file = st.file_uploader(_t("ops_choose_file"), type=["xlsx", "xls", "csv"], key="ops_uploader")
    
    if ops_uploaded_file:
        try:
            # ---------- 1. 文件读取 ----------
            df_ops = None
            file_name = ops_uploaded_file.name
            
            if file_name.endswith(".csv"):
                try:
                    df_ops = pd.read_csv(
                        ops_uploaded_file,
                        encoding='utf-8-sig',
                        engine='python',
                        on_bad_lines='warn'
                    )
                except UnicodeDecodeError:
                    df_ops = pd.read_csv(
                        ops_uploaded_file,
                        encoding='gbk',
                        engine='python',
                        on_bad_lines='warn'
                    )
                except Exception as e:
                    st.error(f"读取 CSV 失败：{e}")
                    st.stop()
                
                if df_ops is not None and len(df_ops.columns) == 1:
                    st.warning("检测到只有 1 列，可能分隔符不是逗号，尝试自动检测分隔符...")
                    ops_uploaded_file.seek(0)
                    content = ops_uploaded_file.read().decode('utf-8-sig', errors='replace')
                    import csv
                    sniffer = csv.Sniffer()
                    try:
                        dialect = sniffer.sniff(content[:1024])
                        ops_uploaded_file.seek(0)
                        df_ops = pd.read_csv(
                            ops_uploaded_file,
                            encoding='utf-8-sig',
                            sep=dialect.delimiter,
                            engine='python',
                            on_bad_lines='warn'
                        )
                    except Exception as e:
                        st.error(f"自动检测分隔符失败：{e}")
                        st.stop()
            else:
                df_ops = pd.read_excel(ops_uploaded_file, engine='openpyxl')
            
            # ---------- 2. 显示读取行数 ----------
            st.info(f"📄 从文件中读取了 **{len(df_ops)}** 行数据，共 **{len(df_ops.columns)}** 列")
            st.write(f"列名: {list(df_ops.columns)}")
            
            if len(df_ops) == 0:
                st.error("文件内容为空，请检查文件。")
                st.stop()
            
            # ---------- 2.5 按唯一键去重 ----------
            df_ops = df_ops.drop_duplicates(
                subset=['biz_date', 'station_code', 'station_name', 'class_name', 'is_conso'],
                keep='last'
            )
            st.info(f"📄 去重后有效数据共 **{len(df_ops)}** 行")
            
            # ---------- 3. 列名校验 ----------
            if list(df_ops.columns) != OPS_COLS:
                st.error(_t("ops_invalid_columns"))
                st.write("模板列名：", OPS_COLS)
                st.write("您的列名：", list(df_ops.columns))
                st.stop()
            
            # ---------- 4. 数据类型转换 ----------
            df_ops["biz_date"] = pd.to_datetime(df_ops["biz_date"]).dt.strftime("%Y-%m-%d")
            df_ops["volume"] = pd.to_numeric(df_ops["volume"])
            df_ops["is_conso"] = pd.to_numeric(df_ops["is_conso"])
            
            st.success(_t("ops_validation_passed", count=len(df_ops)))
            st.dataframe(df_ops.head(10), use_container_width=True)
            
            # ---------- 5. 提交按钮 ----------
            if st.button(_t("ops_submit_btn"), use_container_width=True, key="ops_submit_btn"):
                with st.status(_t("ops_uploading_status"), expanded=True) as status:
                    df_ops["上传人"] = user
                    df_ops["上传时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    df_ops["版本号"] = datetime.now().strftime("%Y%m%d") + "V1"
                    
                    success_count, fail_count = save_operations_to_db(df_ops)
                    
                    if fail_count == 0:
                        msg = f"✅ 成功上传 {success_count} 条操作量数据！"
                        st.session_state["ops_upload_success_msg"] = msg
                        status.update(label=f"✅ 成功导入 {success_count} 条数据！", state="complete")
                        st.balloons()
                        st.rerun()
                    else:
                        status.update(label=f"⚠️ 导入完成，成功 {success_count} 条，失败 {fail_count} 条", state="error")
                        st.error(f"❌ 导入失败 {fail_count} 条，请检查数据")
                        st.write(f"👤 上传人：**{user}**")
                        st.write(f"📅 上传时间：**{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}**")
                        st.write(f"📊 总行数：{len(df_ops)}")
                        st.write(f"✅ 成功：{success_count} 条")
                        st.write(f"❌ 失败：{fail_count} 条")
            
            # ---------- 6. 显示成功消息 ----------
            if "ops_upload_success_msg" in st.session_state and st.session_state["ops_upload_success_msg"]:
                st.success(st.session_state["ops_upload_success_msg"])
                if st.button(_t("clear_msg"), key="clear_ops_msg_btn"):
                    st.session_state["ops_upload_success_msg"] = ""
                    st.rerun()
                    
        except Exception as e:
            st.error(_t("ops_read_error", error=str(e)))
            import traceback
            st.code(traceback.format_exc())
    
    st.divider()
    st.subheader(_t("ops_overview_title"))
    st.caption(_t("ops_overview_caption"))
    try:
        ops_df = get_operation_data()
        if len(ops_df) == 0:
            st.info(_t("ops_no_data"))
        else:
            ops_df["年月"] = pd.to_datetime(ops_df["日期"]).dt.strftime("%Y-%m")
            
            st.subheader(_t("filter_conditions"))
            
            region_order = ["CDC", "SP", "South", "South East", "North East", "North East 2", "Middle West", "North", "FM"]
            all_regions = [r for r in region_order if r in REGION_WAREHOUSE_MAPPING.keys()]
            all_sites_in_data = sorted(ops_df["仓库"].unique())
            
            # ---------- 日期范围筛选 ----------
            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            with col_f1:
                start_date_ops = st.date_input(
                    _t("start_date"),
                    value=pd.to_datetime(ops_df["日期"]).min(),
                    key="ops_start_date"
                )
            with col_f2:
                end_date_ops = st.date_input(
                    _t("end_date"),
                    value=pd.to_datetime(ops_df["日期"]).max(),
                    key="ops_end_date"
                )
            with col_f3:
                selected_region_ops = st.selectbox(
                    _t("select_region"),
                    [_t("all_regions")] + all_regions,
                    key="ops_overview_region"
                )
            with col_f4:
                if selected_region_ops == _t("all_regions"):
                    available_sites = all_sites_in_data
                else:
                    region_sites = REGION_WAREHOUSE_MAPPING.get(selected_region_ops, [])
                    available_sites = [site for site in region_sites if site in all_sites_in_data]
                
                selected_site_ops = st.selectbox(
                    _t("select_site"),
                    [_t("all_sites")] + available_sites,
                    key="ops_overview_site"
                )
            
            # ---------- 数据过滤 ----------
            ops_filtered = ops_df.copy()
            ops_filtered["日期"] = pd.to_datetime(ops_filtered["日期"])
            ops_filtered = ops_filtered[
                (ops_filtered["日期"] >= pd.to_datetime(start_date_ops)) & 
                (ops_filtered["日期"] <= pd.to_datetime(end_date_ops))
            ]
            if selected_region_ops != _t("all_regions"):
                region_sites = REGION_WAREHOUSE_MAPPING.get(selected_region_ops, [])
                ops_filtered = ops_filtered[ops_filtered["仓库"].isin(region_sites)]
            if selected_site_ops != _t("all_sites"):
                ops_filtered = ops_filtered[ops_filtered["仓库"] == selected_site_ops]
            
            # ---------- 指标卡片 ----------
            total_ops_records = len(ops_filtered)
            total_ops_volume = int(ops_filtered["操作量"].sum()) if total_ops_records > 0 else 0
            total_ops_sites = ops_filtered["仓库"].nunique() if total_ops_records > 0 else 0
            total_ops_days = ops_filtered["日期"].nunique() if total_ops_records > 0 else 0
            daily_avg_volume = total_ops_volume / total_ops_days if total_ops_days > 0 else 0
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric(_t("total_volume"), f"{total_ops_volume:,}")
            col2.metric(_t("avg_daily_volume"), f"{daily_avg_volume:,.0f}")
            col3.metric(_t("sites_count"), total_ops_sites)
            col4.metric(_t("total_records"), total_ops_records)
            
            st.divider()
            
            # ---------- 操作量明细（直接展示原始数据） ----------
            st.subheader(_t("ops_detail"))
            if total_ops_records > 0:
                # 复制数据，避免修改原始数据
                display_df = ops_filtered.copy()
                
                # 确保 is_conso 列存在（如果不存在，添加默认值0）
                if "is_conso" not in display_df.columns:
                    display_df["is_conso"] = 0
                
                # 准备要显示的列：日期、站点、班次、是否集包、操作量
                display_df[_t("ops_col_date")] = pd.to_datetime(display_df["日期"]).dt.strftime("%Y-%m-%d")
                display_df[_t("ops_col_site")] = display_df["仓库"]
                display_df[_t("ops_col_shift")] = display_df["班次"]
                display_df[_t("ops_col_is_conso")] = display_df["is_conso"]
                display_df[_t("ops_col_volume")] = display_df["操作量"]
                
                # 按指定顺序选择列
                display_df = display_df[[_t("ops_col_date"), _t("ops_col_site"), _t("ops_col_shift"), _t("ops_col_is_conso"), _t("ops_col_volume")]]
                
                st.dataframe(display_df, use_container_width=True)
            else:
                st.info(_t("ops_no_data"))
            
            # ---------- 导出 ----------
            st.subheader(_t("ops_export"))
            ops_csv = ops_filtered.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                label=_t("ops_export_csv"),
                data=ops_csv,
                file_name=f"操作量数据_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                key="ops_export_btn"
            )
    except Exception as e:
        st.warning(_t("ops_read_data_error", error=str(e)))

# ===================== Tab 5: 价卡配置 =====================
else:
    st.title(_t("price_title"))
    if not is_admin:
        st.warning(_t("price_admin_only"))
        price_df = get_price_card_data()
        if len(price_df) > 0:
            st.subheader(_t("price_current_list"))
            # 使用新的列列表（不含区域）
            display_cols = [col for col in PRICE_CARD_COLS if col in price_df.columns]
            rename_map = {
                "仓库": _t("col_warehouse_name"),
                "供应商": _t("col_supplier"),
                "班次": _t("col_shift"),
                "长期工_日结工": _t("col_worker_type"),
                "周日_非周日": _t("col_sunday"),
                "单价": _t("col_unit_price"),
                "生效时间": _t("col_effective_date"),
                "失效时间": _t("col_expiry_date")
            }
            df_display = price_df[display_cols].rename(columns=rename_map)
            if _t("col_worker_type") in df_display.columns:
                df_display[_t("col_worker_type")] = df_display[_t("col_worker_type")].map(
                    lambda x: _t("worker_long") if x == "长期工" else (_t("worker_daily") if x == "日结工" else x)
                )
            st.dataframe(df_display, use_container_width=True)
        else:
            st.info(_t("price_no_config"))
        st.stop()
    
    st.success(_t("price_admin_mode"))
    price_df = get_price_card_data()
    if len(price_df) > 0:
        latest_version = price_df["版本号"].max()
        st.info(_t("price_current_version", version=latest_version))
    else:
        st.info(_t("price_no_version"))
    
    st.subheader(_t("price_download_template"))
    # 模板去掉"区域"列
    price_template_df = pd.DataFrame({
        "仓库": ["TP IMN"],
        "供应商": ["Enfok"],
        "班次": ["T1"],
        "长期工_日结工": ["长期工"],
        "周日_非周日": ["周日"],
        "单价": [178],
        "生效时间": ["2026-04-01"],
        "失效时间": ["2099-12-31"]
    })
    # 确保列顺序与 PRICE_CARD_COLS 一致
    price_template_df = price_template_df[PRICE_CARD_COLS]
    price_output = BytesIO()
    with pd.ExcelWriter(price_output, engine="openpyxl") as writer:
        price_template_df.to_excel(writer, index=False, sheet_name="价卡配置")
    price_template_bytes = price_output.getvalue()
    col_t1, col_t2 = st.columns([1, 3])
    with col_t1:
        st.download_button(
            label=_t("price_download_btn"),
            data=price_template_bytes,
            file_name="价卡配置模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="price_download_btn"
        )
    # 更新提示信息（去掉"区域"）
    st.caption(_t("price_template_cols"))
    st.divider()

    st.subheader(_t("price_upload_instruction"))
    st.caption(_t("price_upload_caption"))
    with st.form("upload_price_version_form"):
        uploaded_price_file = st.file_uploader(_t("price_choose_file"), type=["xlsx", "xls", "csv"], key="price_uploader")
        submit_price = st.form_submit_button(_t("price_submit_btn"))
    if submit_price:
        if uploaded_price_file is None:
            st.error(_t("price_missing_file"))
        else:
            try:
                df_price = pd.read_excel(uploaded_price_file) if not uploaded_price_file.name.endswith(".csv") else pd.read_csv(uploaded_price_file)
                # 使用新的 PRICE_CARD_COLS 校验列名
                if list(df_price.columns) != PRICE_CARD_COLS:
                    st.error(_t("price_invalid_columns"))
                    st.write("模板列名：", PRICE_CARD_COLS)
                    st.write("您的列名：", list(df_price.columns))
                    st.stop()
                try:
                    df_price["生效时间"] = pd.to_datetime(df_price["生效时间"]).dt.strftime("%Y-%m-%d")
                    df_price["失效时间"] = pd.to_datetime(df_price["失效时间"]).dt.strftime("%Y-%m-%d")
                except Exception as e:
                    st.error(_t("price_date_error", error=str(e)))
                    st.stop()
                try:
                    df_price["单价"] = pd.to_numeric(df_price["单价"])
                except Exception as e:
                    st.error(_t("price_price_error", error=str(e)))
                    st.stop()
                version_name = datetime.now().strftime("%Y%m%d%H%M%S")
                df_price["上传人"] = user
                df_price["上传时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                df_price["版本号"] = version_name
                success_count, fail_count = save_price_card_to_db(df_price)
                if fail_count == 0:
                    st.success(_t("price_upload_success", version=version_name, count=len(df_price)))
                    st.balloons()
                    st.rerun()
                else:
                    st.error(_t("price_upload_error"))
            except Exception as e:
                st.error(_t("price_read_error", error=str(e)))
    
    st.divider()
    st.subheader(_t("price_current_list"))
    price_df = get_price_card_data()
    if len(price_df) > 0:
        latest_version = price_df["版本号"].max()
        latest_df = price_df[price_df["版本号"] == latest_version]
        uploader_name = latest_df['上传人'].iloc[0] if '上传人' in latest_df.columns else '-'
        st.caption(_t("price_current_version_detail", version=latest_version, uploader=uploader_name))
        # 使用新的列列表（不含区域）
        display_cols = [col for col in PRICE_CARD_COLS if col in latest_df.columns]
        rename_map = {
            "仓库": _t("col_warehouse_name"),
            "供应商": _t("col_supplier"),
            "班次": _t("col_shift"),
            "长期工_日结工": _t("col_worker_type"),
            "周日_非周日": _t("col_sunday"),
            "单价": _t("col_unit_price"),
            "生效时间": _t("col_effective_date"),
            "失效时间": _t("col_expiry_date")
        }
        df_display = latest_df[display_cols].rename(columns=rename_map)
        if _t("col_worker_type") in df_display.columns:
            df_display[_t("col_worker_type")] = df_display[_t("col_worker_type")].map(
                lambda x: _t("worker_long") if x == "长期工" else (_t("worker_daily") if x == "日结工" else x)
            )
        st.dataframe(df_display, use_container_width=True)
        
        st.subheader(_t("price_export_current"))
        price_csv = latest_df[display_cols].to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            label=_t("price_export_btn", version=latest_version),
            data=price_csv,
            file_name=f"价卡配置_{latest_version}_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            key="price_export_btn"
        )
        with st.expander(_t("price_history_expander")):
            # 修改统计：使用 "仓库" 计数代替 "区域"
            version_summary = price_df.groupby("版本号").agg({
                "上传时间": "first",
                "仓库": "count"   # 改为仓库计数
            }).reset_index()
            version_summary.columns = [_t("price_history_version"), _t("price_history_upload_time"), _t("price_history_records")]
            version_summary = version_summary.sort_values(_t("price_history_upload_time"), ascending=False)
            st.dataframe(version_summary, use_container_width=True)
            selected_version = st.selectbox(
                _t("price_select_version"),
                options=version_summary[_t("price_history_version")].tolist(),
                key="price_ver_select"
            )
            if selected_version:
                version_detail = price_df[price_df["版本号"] == selected_version]
                detail_display = version_detail[display_cols].rename(columns=rename_map)
                if _t("col_worker_type") in detail_display.columns:
                    detail_display[_t("col_worker_type")] = detail_display[_t("col_worker_type")].map(
                        lambda x: _t("worker_long") if x == "长期工" else (_t("worker_daily") if x == "日结工" else x)
                    )
                st.dataframe(detail_display, use_container_width=True)
                detail_csv = version_detail[display_cols].to_csv(index=False).encode("utf-8-sig")
                st.download_button(
                    label=_t("price_export_version_btn", version=selected_version),
                    data=detail_csv,
                    file_name=f"价卡配置_{selected_version}.csv",
                    mime="text/csv",
                    key=f"price_export_ver_{selected_version}"
                )
    else:
        st.info(_t("price_no_config"))

# ========== 底部信息 ==========
role_text = _t("user_role_admin") if is_admin else _t("user_role_user")
st.caption(_t("bottom_info", user=user, role=role_text))
